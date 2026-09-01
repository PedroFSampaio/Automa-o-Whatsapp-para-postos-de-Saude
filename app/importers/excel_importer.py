from pathlib import Path

from .csv_importer import _clean_rows


def read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("Para importar XLSX, instale as dependencias com: pip install -r requirements.txt") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(values)]
    except StopIteration:
        return []
    rows = (dict(zip(headers, values_row)) for values_row in values)
    return _clean_rows(rows)
